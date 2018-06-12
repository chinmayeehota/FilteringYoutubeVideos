from apiclient.discovery import build
from apiclient.errors import HttpError
from oauth2client.tools import argparser
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pickle
import sys

# Set DEVELOPER_KEY to the API key value from the APIs & auth > Registered apps
# tab of
#   https://cloud.google.com/console
# Please ensure that you have enabled the YouTube Data API for your project.
DEVELOPER_KEY = "AIzaSyCm5PTLUvAZX0pmaD7N7XtFxKd9szpARSc"
YOUTUBE_API_SERVICE_NAME = "youtube"
YOUTUBE_API_VERSION = "v3"

def youtube_search():
    youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION,
                    developerKey=DEVELOPER_KEY)
        
    # Call the search.list method to retrieve results matching the specified
    # query terms.
    searchList = ['Sultan full movie','The Martian full movie','Inside Out full movie','Jurassic World full movie','Minions full movie','Insurgent full movie','The Lobster full movie','Lost River full movie','Little Boy full movie','Maggie full movie','Aloha full movie','Consussion full movie','The Revenant full movie','Black Mass full movie','Pixels full movie']
    finalDict = {}
    for searchItem in searchList:
        search_response = youtube.search().list(q=searchItem,part="id,snippet",maxResults=15).execute()
        #print(search_response)
        videos = []
        video_title = []
        playlists = []
                        
        # Add each result to the appropriate list, and then display the lists of
        # matching videos, channels, and playlists.
        for search_result in search_response.get("items", []):
            if search_result["id"]["kind"] == "youtube#video":
                videos.append(search_result["id"]["videoId"])
                video_title.append(search_result["snippet"]["title"])
        videoIdCommentDict = {}
        for i,video_id in enumerate(videos):
            vid_stats = youtube.videos().list(
                                              part="statistics",
                                              id=video_id
                                              ).execute()

            comment_count = vid_stats.get("items")[0].get("statistics").get("commentCount")
            if comment_count != None:
                try:
                    comments = get_comment_threads(youtube, video_id)
                    videoIdCommentDict.update({(video_id,video_title[i]):comments})
                except Exception:
                    print("Exception cought")
                    continue
        finalDict[searchItem] = videoIdCommentDict
    return finalDict



# Call the API's commentThreads.list method to list the existing comment threads.
def get_comment_threads(youtube, video_id):
    comments = []
    results = youtube.commentThreads().list(
                                            part="snippet",
                                            videoId=video_id,
                                            textFormat="plainText",maxResults=20
                                            ).execute()
        
    for item in results["items"]:
        parentId = item["id"]
        comment = item["snippet"]["topLevelComment"]
        parenttext = comment["snippet"]["textDisplay"]
        parenttupple = {}
        parenttupple['0'] = parentId
        parenttupple['1'] = 'none'
        parenttupple['2'] = parenttext
        comments.append(parenttupple)
        childresults = youtube.comments().list(
                                          part="snippet",
                                          parentId=parentId,
                                          textFormat="plainText"
                                          ).execute()
           
         
        for citem in childresults["items"]:
            child_id = citem["id"]
            text = citem["snippet"]["textDisplay"]
            childtupple = {}
            childtupple['0'] = child_id
            childtupple['1'] = parentId
            childtupple['2'] = text
            comments.append(childtupple)
    return comments

def writeUnLabeledDataToExcel(finalDict):
    wb = Workbook()
    dest_filename = 'data.xlsx'
    ws1 = wb.active
    ws1.title = "range names"
    for searchkey,searchValue in finalDict.items():
        for videoDetails,comments in searchValue.items():
            videoId = videoDetails[0]
            videoTitle = videoDetails[1]
            for row in range(len(comments)):
                ws1.append([searchkey,videoId,videoTitle,comments[row]['0'],comments[row]['1'],comments[row]['2']])
    wb.save(filename = dest_filename)

def readLabeledData():
    wb = load_workbook(filename = 'to_tag_youtube_data.xlsx')
    #sheet = wb['final Data']
    sheet = wb['Two class']
    maxRow = sheet.max_row
    maxColumn = sheet.max_column
    finalDict = {}
    for row in range(2,maxRow+1):
        videoId = sheet.cell(row=row, column=2).value
        
        if videoId in finalDict:
            dict = finalDict[videoId]
            commentlist = dict['comments']
            commentlist.append(sheet.cell(row=row, column=6).value)
            dict['comments'] = commentlist
            finalDict[videoId] = dict
        else:
            videoId = sheet.cell(row=row, column=2).value
            comments = []
            searchQuery = sheet.cell(row=row, column=1).value
            videoTitle = sheet.cell(row=row, column=3).value
            comments.append(sheet.cell(row=row, column=6).value)
            relevance = sheet.cell(row=row, column=7).value
            dict = {}
            dict['comments'] = comments
            dict['searchQuery'] = searchQuery
            dict['videoTitle'] = videoTitle
            dict['relevance'] = relevance
            finalDict[videoId] = dict
    return finalDict

if __name__ == "__main__":
    #finalDict = read_xl()
    #pickle.dump(finalDict, open('actualData.pkl', 'wb'))
    if len(sys.argv) < 2:
        print("Need a command line arguement\n0 : To Collect Unlabeled Data from YouTube and write to excel.\n1 : To dump the Labeled data to pickle file\n")
        sys.exit()
    print(sys.argv[1])
    if int(sys.argv[1]) == 0: 
        try:
            finalDict = youtube_search()
        except HttpError as e:
            print ("An HTTP error %d occurred:\n%s" % (e.resp.status, e.content))
        writeUnLabeledDataToExcel(finalDict)
    elif int(sys.argv[1]) == 1:
        finalDict = readLabeledData()
        pickle.dump(finalDict, open('actualData.pkl', 'wb'))
    else:
        print("Incorrect arguement.")
